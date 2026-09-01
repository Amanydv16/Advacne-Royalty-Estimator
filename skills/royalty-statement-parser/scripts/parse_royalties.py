#!/usr/bin/env python3
"""
parse_royalties.py -- turn one or more mapped royalty statements into a single
month/year/amount(EUR) CSV, with an audit trail.
 
You supply a config that says, per file, *which field is the net royalty* and
*which field is the period*. This script does everything after that decision:
exact Decimal arithmetic, period normalisation, currency conversion, grouping,
reconciliation and warnings. It performs no rounding of source values and never
invents a month that is not present in the input.
 
Usage:
    python3 parse_royalties.py --config config.json --out royalties_eur.csv \
                               [--audit audit.md] [--decimals N] [--month-format number|name]
 
Config schema (see references/config-schema.md for the annotated version):
 
{
  "eur_rates": {"USD": "1.17"},          // units of that currency per 1 EUR
  "period_basis": "usage",               // usage | reporting  (documentation only)
  "sources": [
    {
      "label": "DistroKid",
      "file": "results.csv",
      "mode": "rows",                    // one line item per row
      "sheet": "Sheet1",                 // spreadsheets only
      "header_row": 1,                   // 1-based
      "period_column": "Sale Month",
      "amount_column": "Earnings (USD)",
      "currency": "USD",                 // or {"column": "Currency"}
      "filters": [{"column": "Type", "op": "not_equals", "value": "Advance"}]
    },
    {
      "label": "Black 17 Media",
      "file": "statement.xlsx",
      "mode": "cells",                   // summary statement: read labelled cells
      "sheet": "Statement",
      "currency": "USD",
      "entries": [{"period": "Sept 2022", "amount_cell": "B16",
                   "note": "Current Period Royalties"}]
    },
    {
      "label": "Scanned PDF",
      "file": "statement.pdf",
      "mode": "literal",                 // last resort: transcribed figures
      "currency": "GBP",
      "entries": [{"period": "2024-03", "amount": "1234.56",
                   "note": "Net payable, page 1"}]
    }
  ]
}
"""
import argparse
import csv
import json
import os
import re
import sys
from collections import OrderedDict, defaultdict
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation, getcontext
 
getcontext().prec = 34  # plenty for statement-sized figures; we never round inputs
 
 
def dstr(dec):
    """Plain-decimal string, never scientific notation, never rounded.
    Only meaningless trailing zeros are trimmed -- the value is unchanged."""
    s = format(dec, "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s or "0"
 
MONTHS = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9, "oct": 10,
    "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}
MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]
 
CURRENCY_SYMBOL = {"$": "USD", "€": "EUR", "£": "GBP", "¥": "JPY"}
NUMERIC_RE = re.compile(r"^\s*[-+(]?\s*[\$€£¥]?\s*[\d,.\s']*\d\s*\)?\s*$")
 
 
class ParseProblem(Exception):
    pass
 
 
# --------------------------------------------------------------------------
# value parsing
# --------------------------------------------------------------------------
 
def to_decimal(raw, where=""):
    """Exact Decimal from a cell value. Raises rather than guessing."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        raise ParseProblem(f"empty amount at {where}")
    if isinstance(raw, bool):
        raise ParseProblem(f"boolean is not an amount at {where}")
    if isinstance(raw, int):
        return Decimal(raw)
    if isinstance(raw, float):
        # str() of a float is its shortest round-tripping form, i.e. exactly the
        # digits the spreadsheet stored.
        return Decimal(str(raw))
    s = str(raw).strip()
    if not NUMERIC_RE.match(s):
        raise ParseProblem(f"not a number: {raw!r} at {where}")
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    s = re.sub(r"[\$€£¥\s']", "", s)
    if s.endswith("-"):
        neg, s = True, s[:-1]
    if "," in s and "." in s:
        s = s.replace(",", "") if s.rindex(".") > s.rindex(",") else s.replace(".", "").replace(",", ".")
    elif "," in s:
        parts = s.split(",")
        s = s.replace(",", ".") if len(parts) == 2 and len(parts[1]) != 3 else s.replace(",", "")
    try:
        d = Decimal(s)
    except InvalidOperation:
        raise ParseProblem(f"not a number: {raw!r} at {where}")
    return -d if neg else d
 
 
def normalise_year(y):
    y = int(y)
    if y >= 100:
        return y
    return 2000 + y if y <= 69 else 1900 + y
 
 
def parse_period(raw, where=""):
    """Return (year, month) or raise. Accepts the period formats distributors
    actually ship. Quarterly/annual buckets are rejected on purpose: splitting
    them into months would mean inventing numbers."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        raise ParseProblem(f"empty period at {where}")
    if isinstance(raw, (datetime, date)):
        return raw.year, raw.month
    s = str(raw).strip()
 
    if re.search(r"\bq[1-4]\b|\bquarter\b|\bh[12]\b|\bhalf\b|\bfy\b|\bannual\b|\byear[- ]end\b", s, re.I):
        raise ParseProblem(
            f"period {raw!r} at {where} covers more than one month; splitting it "
            f"would fabricate monthly figures. Resolve with the user."
        )
 
    m = re.match(r"^(\d{4})[-/.](\d{1,2})(?:[-/.]\d{1,2})?$", s)          # 2026-04, 2026/04/15
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^(\d{1,2})[-/.](\d{4})$", s)                            # 04/2026
    if m:
        return int(m.group(2)), int(m.group(1))
    m = re.match(r"^(\d{4})(\d{2})$", s)                                   # 202604
    if m and 1 <= int(m.group(2)) <= 12:
        return int(m.group(1)), int(m.group(2))
 
    m = re.match(r"^([A-Za-z]{3,9})\.?[-/ ]+(\d{2,4})$", s)                # April 2026, APR-26
    if m and m.group(1).lower() in MONTHS:
        return normalise_year(m.group(2)), MONTHS[m.group(1).lower()]
    m = re.match(r"^(\d{2,4})[-/ ]+([A-Za-z]{3,9})\.?$", s)                # 2026-Apr
    if m and m.group(2).lower() in MONTHS:
        return normalise_year(m.group(1)), MONTHS[m.group(2).lower()]
 
    m = re.search(r"([A-Za-z]{3,9})\.?,?\s+(\d{4})", s)                    # "for Sept 2022"
    if m and m.group(1).lower() in MONTHS:
        return int(m.group(2)), MONTHS[m.group(1).lower()]
 
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{2,4})$", s)            # 15/04/2026 (day first)
    if m:
        d1, d2, y = int(m.group(1)), int(m.group(2)), normalise_year(m.group(3))
        if d1 > 12 >= d2:
            return y, d2
        if d2 > 12 >= d1:
            return y, d1
        raise ParseProblem(f"ambiguous date {raw!r} at {where}: cannot tell day from month")
 
    raise ParseProblem(f"unrecognised period {raw!r} at {where}")
 
 
def detect_currency_from_text(text):
    """Pull a currency out of a column header like 'Royalty ($US)' or 'Earnings (USD)'."""
    if not text:
        return None
    t = str(text)
    m = re.search(r"\b(USD|EUR|GBP|JPY|AUD|CAD|CHF|SEK|NOK|DKK|BRL|INR|MXN|ZAR|NZD|PLN)\b", t, re.I)
    if m:
        return m.group(1).upper()
    m = re.search(r"\$\s*US|US\s*\$", t, re.I)
    if m:
        return "USD"
    for sym, code in CURRENCY_SYMBOL.items():
        if sym in t:
            return code
    return None
 
 
# --------------------------------------------------------------------------
# readers
# --------------------------------------------------------------------------
 
def read_delimited(path, header_row):
    with open(path, "rb") as fh:
        head = fh.read(65536)
    encoding = "utf-8-sig" if head.startswith(b"\xef\xbb\xbf") else "utf-8"
    try:
        head.decode(encoding)
    except UnicodeDecodeError:
        encoding = "latin-1"
    sample = head.decode(encoding, errors="replace")
    try:
        delim = csv.Sniffer().sniff(sample[:8000], delimiters=",;\t|").delimiter
    except csv.Error:
        delim = max(",;\t|", key=lambda c: sample[:8000].count(c))
    with open(path, newline="", encoding=encoding) as fh:
        table = list(csv.reader(fh, delimiter=delim))
    idx = (header_row or 1) - 1
    return [str(c) for c in table[idx]], table[idx + 1:]
 
 
def read_sheet(path, sheet, header_row):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    idx = (header_row or 1) - 1
    header = [str(c) if c is not None else "" for c in grid[idx]]
    return header, grid[idx + 1:]
 
 
def read_cell(path, sheet, coord):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    return ws[coord].value
 
 
def column_index(header, name, label):
    if isinstance(name, int):
        return name
    norm = [h.strip().lower() for h in header]
    want = str(name).strip().lower()
    if want in norm:
        return norm.index(want)
    hits = [i for i, h in enumerate(norm) if want in h]
    if len(hits) == 1:
        return hits[0]
    raise ParseProblem(
        f"[{label}] column {name!r} not found (or ambiguous) in header {header}"
    )
 
 
def row_passes(row, header, filters, label):
    for f in filters or []:
        idx = column_index(header, f["column"], label)
        val = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ""
        op, target = f.get("op", "equals"), f.get("value")
        if op == "equals" and val != str(target):
            return False
        if op == "not_equals" and val == str(target):
            return False
        if op == "in" and val not in [str(v) for v in target]:
            return False
        if op == "not_in" and val in [str(v) for v in target]:
            return False
        if op == "contains" and str(target).lower() not in val.lower():
            return False
        if op == "not_contains" and str(target).lower() in val.lower():
            return False
        if op == "non_empty" and not val:
            return False
    return True
 
 
# --------------------------------------------------------------------------
# extraction
# --------------------------------------------------------------------------
 
def resolve_currency(spec, header, amount_idx, row, label, warnings):
    if isinstance(spec, dict) and "column" in spec:
        idx = column_index(header, spec["column"], label)
        code = str(row[idx]).strip().upper() if idx < len(row) and row[idx] else ""
        code = CURRENCY_SYMBOL.get(code, code)
        if not code:
            raise ParseProblem(f"[{label}] blank currency cell")
        return code
    if isinstance(spec, str) and spec:
        return spec.upper()
    inferred = detect_currency_from_text(header[amount_idx]) if amount_idx is not None else None
    if inferred:
        warnings.append(f"[{label}] currency not stated in config; inferred {inferred} "
                        f"from column header {header[amount_idx]!r}. Confirm this.")
        return inferred
    raise ParseProblem(f"[{label}] currency could not be determined; set it in the config")
 
 
def extract_source(src, warnings, skipped):
    """Yield dicts: {year, month, currency, amount, label, origin}."""
    label = src.get("label") or os.path.basename(src["file"])
    path = src["file"]
    mode = src.get("mode", "rows")
    items = []
 
    if mode == "literal":
        cur = (src.get("currency") or "").upper()
        for e in src.get("entries", []):
            y, m = parse_period(e["period"], f"{label} literal entry")
            items.append(dict(year=y, month=m,
                              currency=(e.get("currency") or cur).upper(),
                              amount=to_decimal(e["amount"], label),
                              label=label,
                              origin=e.get("note", "literal")))
        return items
 
    if mode == "cells":
        cur = (src.get("currency") or "").upper()
        for e in src.get("entries", []):
            y, m = parse_period(e["period"], f"{label} cell entry")
            raw = read_cell(path, src.get("sheet"), e["amount_cell"])
            items.append(dict(year=y, month=m,
                              currency=(e.get("currency") or cur).upper(),
                              amount=to_decimal(raw, f"{label}!{e['amount_cell']}"),
                              label=label,
                              origin=f"{e.get('note', 'cell')} @{e['amount_cell']}"))
        return items
 
    # mode == "rows"
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xltx"):
        header, rows = read_sheet(path, src.get("sheet"), src.get("header_row", 1))
    else:
        header, rows = read_delimited(path, src.get("header_row", 1))

    amount_idx = column_index(header, src["amount_column"], label)
    fixed_period = src.get("period")
    year_col = src.get("year_column")
    month_col = src.get("month_column")
    year_idx = column_index(header, year_col, label) if year_col else None
    month_idx = column_index(header, month_col, label) if month_col else None
    period_idx = None if (fixed_period or (year_idx is not None and month_idx is not None)) else column_index(header, src.get("period_column"), label)

    for n, row in enumerate(rows, start=(src.get("header_row", 1) + 1)):
        if not any(str(c).strip() for c in row if c is not None):
            continue
        try:
            if not row_passes(row, header, src.get("filters"), label):
                continue
            raw_amount = row[amount_idx] if amount_idx < len(row) else None
            if raw_amount is None or str(raw_amount).strip() == "":
                continue  # blank amount cell contributes nothing; not an error
            amount = to_decimal(raw_amount, f"{label} row {n}")

            if year_idx is not None and month_idx is not None:
                raw_y = row[year_idx] if year_idx < len(row) else None
                raw_m = row[month_idx] if month_idx < len(row) else None
                if raw_y is None or raw_m is None or str(raw_y).strip() == "" or str(raw_m).strip() == "":
                    raise ParseProblem(f"empty year or month cell at {label} row {n}")
                y = normalise_year(raw_y)
                m_str = str(raw_m).strip().lower()
                m = MONTHS[m_str] if m_str in MONTHS else int(m_str)
            else:
                y, m = parse_period(fixed_period if fixed_period else row[period_idx],
                                    f"{label} row {n}")

            currency = resolve_currency(src.get("currency"), header, amount_idx, row, label, warnings)
            items.append(dict(year=y, month=m, currency=currency, amount=amount,
                              label=label, origin=f"row {n}"))
        except ParseProblem as exc:
            skipped.append(f"[{label}] row {n}: {exc}")
    return items
 
 
# --------------------------------------------------------------------------
# aggregation + output
# --------------------------------------------------------------------------
 
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--audit", default=None)
    ap.add_argument("--decimals", type=int, default=None,
                    help="round the EUR column to N places for presentation. "
                         "Omit to keep full precision (the default).")
    ap.add_argument("--month-format", choices=["number", "name"], default="number")
    args = ap.parse_args()
 
    with open(args.config, encoding="utf-8") as fh:
        cfg = json.load(fh)
 
    rates = {k.upper(): Decimal(str(v)) for k, v in (cfg.get("eur_rates") or {}).items()}
    rates.setdefault("EUR", Decimal("1"))
 
    warnings, skipped, items = [], [], []
    for src in cfg["sources"]:
        try:
            items.extend(extract_source(src, warnings, skipped))
        except (ParseProblem, KeyError, FileNotFoundError) as exc:
            warnings.append(f"[{src.get('label', src.get('file'))}] SOURCE FAILED: {exc}")
 
    if not items:
        print("No line items extracted. Check the config mapping.", file=sys.stderr)
 
    # group by (year, month, currency) and sum exactly, then convert once per
    # bucket -- summing before dividing avoids compounding division error.
    buckets = defaultdict(Decimal)
    per_source = defaultdict(Decimal)
    for it in items:
        buckets[(it["year"], it["month"], it["currency"])] += it["amount"]
        per_source[(it["label"], it["year"], it["month"], it["currency"])] += it["amount"]
 
    missing_rates = sorted({c for (_, _, c) in buckets if c not in rates})
    for c in missing_rates:
        warnings.append(
            f"No EUR rate supplied for {c}. Those months are written with a null "
            f"amount rather than an invented conversion. Add it to eur_rates."
        )
 
    months = defaultdict(lambda: {"eur": Decimal(0), "convertible": False, "unconverted": []})
    for (y, m, cur), total in buckets.items():
        entry = months[(y, m)]
        if cur in rates:
            entry["eur"] += total / rates[cur]
            entry["convertible"] = True
        else:
            entry["unconverted"].append((cur, total))
 
    # A month that mixes convertible and unconvertible currencies must not be
    # written as a partial total -- that would understate it while looking valid.
    for (y, m), entry in months.items():
        if entry["unconverted"]:
            if entry["convertible"]:
                warnings.append(
                    f"{y}-{m:02d} contains amounts in "
                    f"{', '.join(c for c, _ in entry['unconverted'])} with no EUR rate. "
                    f"Writing null instead of a partial total that would understate the month."
                )
            entry["convertible"] = False
 
    def fmt(dec):
        if args.decimals is not None:
            return format(dec.quantize(Decimal(1).scaleb(-args.decimals),
                                       rounding=ROUND_HALF_UP), "f")
        return dstr(dec)
 
    ordered = OrderedDict(sorted(months.items()))
    os.makedirs(os.path.dirname(os.path.abspath(args.out)) or ".", exist_ok=True)
    with open(args.out, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["month", "year", "amount"])
        for (y, m), entry in ordered.items():
            month_cell = MONTH_NAMES[m] if args.month_format == "name" else m
            amount_cell = fmt(entry["eur"]) if entry["convertible"] else ""
            w.writerow([month_cell, y, amount_cell])
 
    # ---- audit -----------------------------------------------------------
    lines = ["# Royalty parse audit", ""]
    lines.append(f"- Output: `{args.out}`")
    lines.append(f"- Period basis: **{cfg.get('period_basis', 'unspecified')}** "
                 f"(the month each amount is attributed to)")
    lines.append(f"- EUR rates used (units per 1 EUR): "
                 f"{ {k: str(v) for k, v in rates.items()} }")
    lines.append(f"- Line items extracted: {len(items)}  |  months emitted: {len(ordered)}")
    lines.append(f"- EUR column rounding: "
                 f"{'none (full precision)' if args.decimals is None else str(args.decimals) + ' dp'}")
    lines.append("")
 
    lines += ["## Per-source subtotals (source currency, unrounded)", "",
              "| Source | Year | Month | Currency | Source total | EUR |",
              "|---|---|---|---|---|---|"]
    for (label, y, m, cur), total in sorted(per_source.items()):
        eur = dstr(total / rates[cur]) if cur in rates else "NULL"
        lines.append(f"| {label} | {y} | {m:02d} | {cur} | {dstr(total)} | {eur} |")
 
    lines += ["", "## Monthly output", "", "| Year | Month | EUR amount |", "|---|---|---|"]
    for (y, m), entry in ordered.items():
        lines.append(f"| {y} | {m:02d} | "
                     f"{fmt(entry['eur']) if entry['convertible'] else 'NULL (no rate)'} |")
 
    # reconciliation: EUR total re-multiplied by each rate must return the source total
    lines += ["", "## Reconciliation", ""]
    by_cur = defaultdict(Decimal)
    for (y, m, cur), total in buckets.items():
        by_cur[cur] += total
    for cur, total in sorted(by_cur.items()):
        back = (total / rates[cur]) * rates[cur] if cur in rates else None
        ok = "OK" if back is not None and abs(back - total) < Decimal("1e-20") else (
            "N/A (no rate)" if back is None else "CHECK")
        lines.append(f"- {cur}: source total {dstr(total)} -> "
                     f"EUR {dstr(total / rates[cur]) if cur in rates else 'NULL'} "
                     f"[round-trip {ok}]")
 
    gaps = []
    if ordered:
        keys = list(ordered)
        y, m = keys[0]
        while (y, m) != keys[-1]:
            m += 1
            if m == 13:
                y, m = y + 1, 1
            if (y, m) not in ordered and (y, m) != keys[-1]:
                gaps.append(f"{y}-{m:02d}")
    if gaps:
        lines += ["", "## Calendar gaps (deliberately NOT emitted)", "",
                  "These months fall inside the covered range but appear in no source "
                  "document, so no row was written for them. They are absent, not zero.",
                  "", ", ".join(gaps[:24]) + (f", ... (+{len(gaps) - 24} more)" if len(gaps) > 24 else "")]
 
    if warnings:
        lines += ["", "## Warnings", ""] + [f"- {w}" for w in warnings]
    if skipped:
        lines += ["", f"## Skipped rows ({len(skipped)})", ""] + [f"- {s}" for s in skipped[:80]]
        if len(skipped) > 80:
            lines += [f"- ... {len(skipped) - 80} more"]
 
    audit_text = "\n".join(lines) + "\n"
    if args.audit:
        with open(args.audit, "w", encoding="utf-8") as fh:
            fh.write(audit_text)
 
    print(audit_text)
    if warnings or skipped:
        print(f"\n>>> {len(warnings)} warning(s), {len(skipped)} skipped row(s). "
              f"Review before sending the CSV to the user.", file=sys.stderr)
 
 
if __name__ == "__main__":
    main()

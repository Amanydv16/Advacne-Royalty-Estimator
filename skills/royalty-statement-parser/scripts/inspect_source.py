#!/usr/bin/env python3
"""
inspect_source.py -- profile a royalty statement so you can decide which field
holds the artist's net royalty and which field holds the period.
 
Prints structure only. It never guesses a mapping for you and never writes output.
 
Usage:
    python3 inspect_source.py <file> [<file> ...] [--rows N] [--full]
 
    --rows N   number of sample data rows to print (default 5)
    --full     for spreadsheets, dump every non-empty cell with its coordinate
               (use for summary-style statements; automatic for sheets < 400 cells)
"""
import argparse
import csv
import io
import json
import os
import re
import sys
from collections import Counter
from decimal import Decimal, InvalidOperation
 
MAX_AUTO_FULL_CELLS = 400
NUMERIC_RE = re.compile(r"^\s*[-+(]?\s*[\$€£¥]?\s*[\d,.\s']*\d\s*\)?\s*[%]?\s*$")
 
 
def to_decimal(raw):
    """Parse a spreadsheet/CSV value into Decimal, or None if not numeric.
 
    Handles $1,234.56, (1.234,56), 1 234,56, trailing minus. Returns the value
    unrounded -- callers must never re-round it.
    """
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int,)):
        return Decimal(raw)
    if isinstance(raw, float):
        return Decimal(str(raw))
    s = str(raw).strip()
    if not s or not NUMERIC_RE.match(s):
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    s = re.sub(r"[\$€£¥%\s']", "", s)
    if s.endswith("-"):
        neg, s = True, s[:-1]
    if "," in s and "." in s:
        # whichever separator is last is the decimal separator
        s = s.replace(",", "") if s.rindex(".") > s.rindex(",") else s.replace(".", "").replace(",", ".")
    elif "," in s:
        frac = s.split(",")[-1]
        s = s.replace(",", ".") if len(s.split(",")) == 2 and len(frac) != 3 else s.replace(",", "")
    if s in ("", "-", "+", "."):
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return -d if neg else d
 
 
def sniff_text(path):
    with open(path, "rb") as fh:
        raw = fh.read(200_000)
    encoding = "utf-8-sig" if raw.startswith(b"\xef\xbb\xbf") else "utf-8"
    try:
        raw.decode(encoding)
    except UnicodeDecodeError:
        encoding = "latin-1"
    sample = raw.decode(encoding, errors="replace")
    try:
        delim = csv.Sniffer().sniff(sample[:8000], delimiters=",;\t|").delimiter
    except csv.Error:
        delim = max(",;\t|", key=lambda c: sample[:8000].count(c))
    return encoding, delim
 
 
def profile_columns(header, rows):
    """For each column: fill rate, distinct values (if few), and an exact sum
    when numeric. The sums are the fastest way to tell a net-royalty column from
    a gross / withholding / balance column."""
    out = []
    for i, name in enumerate(header):
        values = [r[i] if i < len(r) else None for r in rows]
        nonempty = [v for v in values if v not in (None, "")]
        nums = [to_decimal(v) for v in nonempty]
        nums = [n for n in nums if n is not None]
        col = {
            "column": name,
            "filled": f"{len(nonempty)}/{len(values)}",
            "samples": [str(v) for v in nonempty[:3]],
        }
        if nums and len(nums) >= max(1, len(nonempty) * 0.9):
            col["numeric"] = True
            col["exact_sum"] = str(sum(nums))
            col["min"] = str(min(nums))
            col["max"] = str(max(nums))
            col["negatives"] = sum(1 for n in nums if n < 0)
        else:
            distinct = Counter(str(v) for v in nonempty)
            col["distinct"] = len(distinct)
            if len(distinct) <= 15:
                col["values"] = dict(distinct.most_common(15))
        out.append(col)
    return out
 
 
def inspect_delimited(path, args):
    encoding, delim = sniff_text(path)
    with open(path, newline="", encoding=encoding) as fh:
        reader = csv.reader(fh, delimiter=delim)
        table = list(reader)
    if not table:
        print("  (empty file)")
        return
    header_idx = 0
    for idx, row in enumerate(table[:20]):
        if sum(1 for c in row if str(c).strip()) >= 3:
            header_idx = idx
            break
    header, rows = table[header_idx], table[header_idx + 1:]
    print(f"  encoding={encoding}  delimiter={delim!r}  header_row={header_idx + 1}  data_rows={len(rows)}")
    print(f"  columns ({len(header)}): {header}")
    print("\n  -- column profile (exact sums, unrounded) --")
    for col in profile_columns(header, rows):
        print("   " + json.dumps(col, ensure_ascii=False))
    print(f"\n  -- first {args.rows} data rows --")
    for row in rows[:args.rows]:
        print("   " + json.dumps(row, ensure_ascii=False))
 
 
def inspect_xlsx(path, args):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    print(f"  sheets: {wb.sheetnames}")
    for ws in wb.worksheets:
        cells = ws.max_row * ws.max_column
        print(f"\n  === sheet {ws.title!r}  dims={ws.dimensions}  rows={ws.max_row} cols={ws.max_column} ===")
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        dense = [r for r in grid if sum(1 for c in r if c not in (None, "", " ")) >= 3]
        looks_tabular = len(dense) >= 8 and len(dense) > 0.5 * max(1, len(grid))
        if args.full or cells <= MAX_AUTO_FULL_CELLS or not looks_tabular:
            print("  -- every non-empty cell (coordinate: value) --")
            print("  NOTE: summary statements put the payable figure in a labelled cell.")
            print("        Reference it by coordinate in the config so the value is read")
            print("        from the file rather than retyped by hand.")
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value not in (None, "", " "):
                        print(f"   {cell.coordinate}: {cell.value!r}")
        else:
            header_idx = next((i for i, r in enumerate(grid)
                               if sum(1 for c in r if c not in (None, "", " ")) >= 3), 0)
            header = [str(c) if c is not None else "" for c in grid[header_idx]]
            rows = grid[header_idx + 1:]
            print(f"  header_row={header_idx + 1}  data_rows={len(rows)}")
            print(f"  columns: {header}")
            print("\n  -- column profile (exact sums, unrounded) --")
            for col in profile_columns(header, rows):
                print("   " + json.dumps(col, ensure_ascii=False, default=str))
            print(f"\n  -- first {args.rows} data rows --")
            for row in rows[:args.rows]:
                print("   " + json.dumps([str(c) if c is not None else "" for c in row], ensure_ascii=False))
 
 
def inspect_pdf(path, args):
    try:
        import pdfplumber
    except ImportError:
        print("  pdfplumber not installed: pip install pdfplumber --break-system-packages")
        return
    with pdfplumber.open(path) as pdf:
        print(f"  pages: {len(pdf.pages)}")
        for pno, page in enumerate(pdf.pages[:5], 1):
            print(f"\n  === page {pno} text ===")
            print("   " + (page.extract_text() or "(no text layer)").replace("\n", "\n   "))
            for tno, table in enumerate(page.extract_tables(), 1):
                print(f"\n  === page {pno} table {tno} ===")
                for row in table[:args.rows + 1]:
                    print("   " + json.dumps(row, ensure_ascii=False))
        if len(pdf.pages) > 5:
            print(f"\n  ({len(pdf.pages) - 5} further pages not shown; rerun on a slice if needed)")
 
 
def inspect_json(path, args):
    with open(path, encoding="utf-8") as fh:
        data = json.load(fh)
 
    def shape(node, depth=0):
        pad = "   " + "  " * depth
        if isinstance(node, dict):
            for k, v in list(node.items())[:25]:
                if isinstance(v, (dict, list)):
                    print(f"{pad}{k}: {type(v).__name__}")
                    if depth < 3:
                        shape(v, depth + 1)
                else:
                    print(f"{pad}{k}: {v!r}")
        elif isinstance(node, list):
            print(f"{pad}[{len(node)} items]")
            if node and depth < 3:
                shape(node[0], depth + 1)
 
    shape(data)
 
 
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="+")
    ap.add_argument("--rows", type=int, default=5)
    ap.add_argument("--full", action="store_true")
    args = ap.parse_args()
 
    for path in args.files:
        print("=" * 78)
        print(f"FILE: {path}  ({os.path.getsize(path):,} bytes)")
        print("=" * 78)
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext in (".xlsx", ".xlsm", ".xltx"):
                inspect_xlsx(path, args)
            elif ext == ".pdf":
                inspect_pdf(path, args)
            elif ext == ".json":
                inspect_json(path, args)
            elif ext == ".xls":
                print("  legacy .xls -- convert first: libreoffice --headless --convert-to xlsx <file>")
            else:
                inspect_delimited(path, args)
        except Exception as exc:  # keep going across a batch of statements
            print(f"  ERROR: {type(exc).__name__}: {exc}")
        print()
 
 
if __name__ == "__main__":
    main()
